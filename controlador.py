
import os

from Modelo.negocio import *
from Modelo.provincia import *
from Modelo.localidad import *

from autowordpress import *

import openpyxl

from funciones import crea_provincia, crea_provincia, obten_nombre_municipio

class Controlador:
    def __init__(self,datos_empresas,datos_municipios,datos_provincias):
        self.datos_empresas = datos_empresas
        self.datos_localidades = datos_municipios
        self.datos_provincias = datos_provincias

        self.tipo_negocio="Empresas de alquiler de maquinaria"
        self.dominio="https://psicosensometrico-ecuador.com/"
        self.color_base="#1F2937"
        self.color_base2="#374151"
        self.color_base3="#F3F4F6"
        self.color_contrast="#111827"
        self.color_contrast2="#9CA3AF"
        self.color_contrast3="#E5E7EB"
        self.color_accent="#F2B900"
        self.color_noton="#7B2C2C"#accent2

        self.excel_empresas="./xslx/empresas.xlsx"
        self.excel_municipios="./xslx/localidades.xlsx"
        self.excel_provincias="./xslx/provincias.xlsx"

        self.tipo_negocio_schema="Store"

        #self.wc=WpConnection(f"{self.dominio}//xmlrpc.php",'hector.arnaus@gmail.com','bolo4o#Eresgay')
        self.wc=WpConnection(f"{self.dominio}//xmlrpc.php",'hector.arnaus@gmail.com','bolo3o,Eresgay')

    def connect(self):
        self.wc.connect()

    def __ultima_fila_real(self,hoja):
        # Recorre desde abajo hacia arriba buscando la primera fila con algún valor no vacío
        for i in range(1,hoja.max_row, 1):
            if hoja.cell(row=i,column=1).value==None:
               return i
        return hoja.max_row
    
    def __limpiar_web(self,web):
        if web!=None:
            if web.find("?utm")!=-1:
                return web[:web.find("?utm")]
        return web
    
    def limpiar_horario(self,horario):
        if horario!=None:
            horario=horario.replace("&quot;",'"').strip('"')
            if not horario.startswith("<li>"):
                horario="<li> "+horario+" </li>"
        return horario
    
    def sluguiza(self,texto):
        texto=texto.strip()
        texto=texto.lower()
        texto=texto.replace("ñ","n")
        texto=texto.replace("á","a")
        texto=texto.replace("ä","a")
        texto=texto.replace("à","a")
        texto=texto.replace("â","a")
        texto=texto.replace("é","e")
        texto=texto.replace("ê","e")
        texto=texto.replace("ë","e")
        texto=texto.replace("è","e")
        texto=texto.replace("í","i")
        texto=texto.replace("ï","i")
        texto=texto.replace("ì","i")
        texto=texto.replace("ó","o")
        texto=texto.replace("ö","o")
        texto=texto.replace("ò","o")
        texto=texto.replace("ú","u")
        texto=texto.replace("ü","u")
        texto=texto.replace("ù","u")
        texto=texto.replace(" ","-")
        texto=texto.replace("&quot;","'")
        texto=texto.replace("&amp;","&")
        texto=texto.replace("&apos;","'")
        texto=texto.replace("|","")
        texto=texto.replace(".","")
        texto=texto.replace(",","")
        texto=texto.replace("*","")
        while texto.find("--")!=-1:
            texto=texto.replace("--","-")

        return texto

    def crea_lista_provincias(self):
        lista=[]
        try:
            datos=openpyxl.load_workbook(self.datos_provincias)
            hoja_activa = datos.active
            fila=2
            while fila<self.__ultima_fila_real(hoja_activa):
                nombre=hoja_activa.cell(row=fila,column=1).value
                if nombre.isupper():
                    nombre=nombre.capitalize()
                actividades=(hoja_activa.cell(row=fila,column=2).value).split(',')
                cabecera=hoja_activa.cell(row=fila,column=3).value
                cuerpo=hoja_activa.cell(row=fila,column=4).value
                lista.append(Provincia(nombre,actividades,cabecera,cuerpo))
                fila+=1
            return lista
        
        except FileNotFoundError:
            print("Error: Archivo no encontrado.")
        except Exception as e:
            print(f"Ocurrió un error: {e}")

    def crea_lista_municipios(self):
        lista=[]
        try:
            datos=openpyxl.load_workbook(self.datos_localidades)
            hoja_activa = datos.active
            fila=2
            while fila<self.__ultima_fila_real(hoja_activa):
                nombre=hoja_activa.cell(row=fila,column=1).value
                if nombre.isupper():
                    nombre=nombre.capitalize()
                provincia=hoja_activa.cell(row=fila,column=2).value
                if provincia.isupper():
                    provincia=provincia.capitalize()
                actividades=(hoja_activa.cell(row=fila,column=4).value).split(',')
                texto=hoja_activa.cell(row=fila,column=5).value
                lista.append(Localidad(nombre,provincia,actividades,texto))
                fila+=1
            return lista

        except FileNotFoundError:
            print("Error: Archivo no encontrado.")
        except Exception as e:
            print(f"Ocurrió un error: {e}")

    def crea_lista_empresas(self):
        lista_empresas=[]
        try:
            datos=openpyxl.load_workbook(self.datos_empresas)
            hoja_activa = datos.active
            fila=1
            while fila<self.__ultima_fila_real(hoja_activa):
                nombre=hoja_activa.cell(row=fila,column=1).value if hoja_activa.cell(row=fila,column=1).value!=None else None
                direccion=hoja_activa.cell(row=fila,column=2).value if hoja_activa.cell(row=fila,column=2).value!=None else None
                CP=hoja_activa.cell(row=fila,column=3).value if hoja_activa.cell(row=fila,column=3).value!=None else None
                ciudad=hoja_activa.cell(row=fila,column=4).value if hoja_activa.cell(row=fila,column=4).value!=None else None
                provincia=hoja_activa.cell(row=fila,column=5).value if hoja_activa.cell(row=fila,column=5).value!=None else None
                telefono=hoja_activa.cell(row=fila,column=6).value if hoja_activa.cell(row=fila,column=6).value!=None else None
                pagina_web=self.__limpiar_web(hoja_activa.cell(row=fila,column=7).value if hoja_activa.cell(row=fila,column=7).value!=None else None)
                actividad=hoja_activa.cell(row=fila,column=8).value if hoja_activa.cell(row=fila,column=8).value!=None else None
                actividades_relacionadas=hoja_activa.cell(row=fila,column=9).value if hoja_activa.cell(row=fila,column=9).value!=None else None
                marcas=hoja_activa.cell(row=fila,column=10).value if hoja_activa.cell(row=fila,column=10).value!=None else None
                descripcion=hoja_activa.cell(row=fila,column=11).value if hoja_activa.cell(row=fila,column=11).value!=None else None
                mapa=hoja_activa.cell(row=fila,column=12).value if hoja_activa.cell(row=fila,column=12).value!=None else None
                imagen=hoja_activa.cell(row=fila,column=13).value if hoja_activa.cell(row=fila,column=13).value!=None else None
                facebook=hoja_activa.cell(row=fila,column=14).value if hoja_activa.cell(row=fila,column=14).value!=None else None
                instagram=hoja_activa.cell(row=fila,column=15).value if hoja_activa.cell(row=fila,column=15).value!=None else None
                x=hoja_activa.cell(row=fila,column=16).value if hoja_activa.cell(row=fila,column=16).value!=None else None
                youtube=hoja_activa.cell(row=fila,column=17).value if hoja_activa.cell(row=fila,column=17).value!=None else None
                horario=self.limpiar_horario(hoja_activa.cell(row=fila,column=18).value if hoja_activa.cell(row=fila,column=18).value!=None else None)
                descripcion_seo=hoja_activa.cell(row=fila,column=19).value  if hoja_activa.cell(row=fila,column=19).value!=None else None
                tagline=hoja_activa.cell(row=fila,column=20).value  if hoja_activa.cell(row=fila,column=20).value!=None else None
                categoria=hoja_activa.cell(row=fila,column=21).value  if hoja_activa.cell(row=fila,column=21).value!=None else None

                nuevo=Negocio(nombre,direccion,CP,ciudad,provincia,telefono,pagina_web,actividad,actividades_relacionadas,marcas,descripcion,mapa,imagen,facebook,instagram,x,youtube,horario,descripcion_seo,tagline,categoria)

                lista_empresas.append(nuevo)
                fila+=1

            return lista_empresas
        except FileNotFoundError:
            print("Error: Archivo no encontrado.")
        except Exception as e:
            print(f"Ocurrió un error: {e}")

    def crea_lista_imagenes_municipio(self,ruta):
        lista_imagenes=[]
        for img in os.listdir(ruta):
            municipio=obten_nombre_municipio(img)    
            if municipio not in lista_imagenes:
                lista_imagenes.append(municipio)
        return lista_imagenes
    
    def crea_lista_municipios_sin_imagen(self,municipios,lista_imagenes):
        lista_municipios_sin_imagen=[]
        for municipio in municipios:
            if municipio.nombre not in lista_imagenes:
                lista_municipios_sin_imagen.append(municipio.nombre)
        return lista_municipios_sin_imagen
    
    def obten_nombre_municipio_a_partir_imagen(self,imagen):
        return imagen.split("_")[0]
    
    def crea_articulo_provincia(self,provincia):
        ruta=os.getcwd()+("/provincia/"+provincia.nombre+".webp")
        wp_img=Image(ruta,f"Descubre todas las {self.tipo_negocio.lower()} en la provincia de {provincia.nombre} ordenadas por orden alfabético")
        wp_img.upload(self.wc)
        wp_article=WpPost(f"{self.tipo_negocio} en la provincia de {provincia.nombre}")
        wp_article.add_element(crea_provincia(provincia,wp_img))
        wp_article.set_slug(self.sluguiza("provincia de "+provincia.nombre))
        wp_article.add_category("provincia")
        self.wc.publica_post(wp_article)