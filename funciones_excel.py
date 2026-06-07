import openpyxl 
from Modelo.provincia import Provincia
from Modelo.negocio import Negocio
from Modelo.localidad import Localidad
from municipio import Municipio
import ast, re



def ultima_fila_real(hoja):
    # Recorre desde abajo hacia arriba buscando la primera fila con algún valor no vacío
    for i in range(1,hoja.max_row, 1):
        if hoja.cell(row=i,column=1).value==None:
           return i
    return hoja.max_row

def crea_lista_provincias(fichero_excel):
    lista=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=2
        while fila<ultima_fila_real(hoja_activa):
            nombre=hoja_activa.cell(row=fila,column=1).value
            if nombre.isupper():
                nombre=nombre.capitalize()
            actividades=hoja_activa.cell(row=fila,column=5).value
            cabecera=hoja_activa.cell(row=fila,column=6).value
            cuerpo=hoja_activa.cell(row=fila,column=7).value
            lista.append(Provincia(nombre,actividades,cabecera,cuerpo))
            fila+=1
        return lista
        
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")



def obten_lista_provincias(fichero_excel):
    lista=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=2
        while fila<ultima_fila_real(hoja_activa):
            provincia=hoja_activa.cell(row=fila,column=5).value
            if provincia.isupper():
                provincia=provincia.capitalize()    
            if provincia not in lista:
                lista.append(provincia)
            fila+=1
        lista.sort()
        return lista
        
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

def obten_lista_municipios(fichero_excel):
    lista=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=2
        while fila<ultima_fila_real(hoja_activa):
            localidad=hoja_activa.cell(row=fila,column=4).value
            if localidad.isupper():
                localidad=localidad.capitalize()    
            if localidad not in lista:
                lista.append(localidad)
            fila+=1
        lista.sort()
        return lista

    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

def obten_lista_municipios_con_provincia(fichero_excel):
    lista=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=2
        while fila<ultima_fila_real(hoja_activa):
            localidad=hoja_activa.cell(row=fila,column=4).value
            if localidad.isupper():
                localidad=localidad.capitalize()    
            if localidad not in lista:
                lista.append([localidad,hoja_activa.cell(row=fila,column=5).value])
            fila+=1
        lista.sort()
        return lista

    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

def obten_provincia_de_municipio(fichero_excel,municipio):
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=2
        while fila<hoja_activa.max_row:
            if hoja_activa.cell(row=fila,column=7).value==municipio:
                return hoja_activa.cell(row=fila,column=8).value
            fila+=1
        return ""
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

def obten_provincia_de_municipio_con_fichero_localidades(fichero_excel,municipio):
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=2
        while fila<hoja_activa.max_row:
            if hoja_activa.cell(row=fila,column=1).value==municipio:
                return hoja_activa.cell(row=fila,column=2).value
            fila+=1
        return ""
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")






def parse_list_text(text):
    """Convierte un texto con formato de lista Python en una lista de strings.
    Ejemplo: "['A', 'B']" -> ['A', 'B']
    Maneja comillas simples/dobles y ofrece fallback si no es literal Python.
    """
    if text is None:
        return []
    text = text.strip()
    if text == "":
        return []
    # intento seguro: evaluar literal Python
    try:
        val = ast.literal_eval(text)
        if isinstance(val, (list, tuple)):
            return [str(x) for x in val]
        if isinstance(val, str):
            # si era una cadena simple, la procesamos más abajo
            text = val
    except Exception:
        pass

    # fallback: extraer items entre comillas
    matches = re.findall(r"(['\"])(.*?)\1", text)
    if matches:
        return [m[1].strip() for m in matches]

    # último recurso: dividir por comas y limpiar
    parts = [p.strip().strip("'\"") for p in re.split(r",\s*", text) if p.strip()]
    return parts

d

def obten_lista_negocios_municipio(fichero_excel,municipio):
    lista_negocios=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=1
        while fila<hoja_activa.max_row:
            if hoja_activa.cell(row=fila,column=4).value==municipio:
                nombre=hoja_activa.cell(row=fila,column=1).value if hoja_activa.cell(row=fila,column=1).value!=None else None
                direccion=hoja_activa.cell(row=fila,column=2).value if hoja_activa.cell(row=fila,column=2).value!=None else None
                CP=hoja_activa.cell(row=fila,column=3).value if hoja_activa.cell(row=fila,column=3).value!=None else None
                provincia=hoja_activa.cell(row=fila,column=5).value if hoja_activa.cell(row=fila,column=5).value!=None else None
                telefono=hoja_activa.cell(row=fila,column=6).value if hoja_activa.cell(row=fila,column=6).value!=None else None
                pagina_web=limpiar_web(hoja_activa.cell(row=fila,column=7).value if hoja_activa.cell(row=fila,column=7).value!=None else None)
                actividad=hoja_activa.cell(row=fila,column=8).value if hoja_activa.cell(row=fila,column=8).value!=None else None
                actividades_relacionadas=hoja_activa.cell(row=fila,column=9).value if hoja_activa.cell(row=fila,column=9).value!=None else None
                marcas=hoja_activa.cell(row=fila,column=10).value if hoja_activa.cell(row=fila,column=10).value!=None else None
                descripcion=hoja_activa.cell(row=fila,column=11).value if hoja_activa.cell(row=fila,column=11).value!=None else None
                mapa=hoja_activa.cell(row=fila,column=12).value if hoja_activa.cell(row=fila,column=12).value!=None else None
                imagen=hoja_activa.cell(row=fila,column=13).value if hoja_activa.cell(row=fila,column=13).value!=None else None
                facebook=hoja_activa.cell(row=fila,column=14).value if hoja_activa.cell(row=fila,column=14).value!=None else None
                instagram=hoja_activa.cell(row=fila,column=15).value if hoja_activa.cell(row=fila,column=15).value!=None else None
                x=hoja_activa.cell(row=fila,column=16).value if hoja_activa.cell(row=fila,column=16).value!=None else None
                youtube=hoja_activa.cell(row=fila,column=17).value if hoja_activa.cell(row=fila,column=17).value!=None else None
                horario=hoja_activa.cell(row=fila,column=18).value if hoja_activa.cell(row=fila,column=18).value!=None else None
                descripcion_seo=hoja_activa.cell(row=fila,column=19).value  if hoja_activa.cell(row=fila,column=19).value!=None else None
                tagline=hoja_activa.cell(row=fila,column=20).value  if hoja_activa.cell(row=fila,column=20).value!=None else None

                nuevo=Negocio(nombre,direccion,CP,municipio,provincia,telefono,pagina_web,actividad,actividades_relacionadas,marcas,descripcion,mapa,imagen,facebook,instagram,x,youtube,horario,descripcion_seo,tagline)
                lista_negocios.append(nuevo)

            fila+=1
                  
        return lista_negocios
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")
               
def obten_municipio_lista(nombre,lista_municipios):
    for municipio in lista_municipios:
        if municipio.nombre==nombre:
            return municipio
    return None

def esta_municipio_en_lista(nombre,lista_municipios):
    for municipio in lista_municipios:
        if municipio.nombre==nombre:
            return True
    return False
def obten_lista_actividades_municipios(fichero_excel):
    lista_municipios=[]
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=1
        while fila<ultima_fila_real(hoja_activa):
            if esta_municipio_en_lista(hoja_activa.cell(row=fila,column=4).value,lista_municipios)==False:
                municipio=Municipio(hoja_activa.cell(row=fila,column=4).value,hoja_activa.cell(row=fila,column=5).value)  
                if hoja_activa.cell(row=fila,column=8)!=None:
                    municipio.anyade_actividad(hoja_activa.cell(row=fila,column=8).value.lower())
                if hoja_activa.cell(row=fila,column=9)!=None:
                    actividades_relacionadas= parse_list_text(hoja_activa.cell(row=fila,column=9).value)
                    for actividad in actividades_relacionadas:
                        municipio.anyade_actividad(actividad.lower())
                lista_municipios.append(municipio)
            else:   
                municipio=obten_municipio_lista(hoja_activa.cell(row=fila,column=4).value,lista_municipios)
                if hoja_activa.cell(row=fila,column=8)!=None:
                    municipio.anyade_actividad(hoja_activa.cell(row=fila,column=8).value.lower())
            fila+=1
        return lista_municipios
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")
               


def obten_ciudades_con_mas_de_un_negocio(fichero_excel):
    
    
    try:
        datos=openpyxl.load_workbook(fichero_excel)
        hoja_activa = datos.active
        fila=1
        lista_ciudades=[]
        while fila<ultima_fila_real(hoja_activa):
           
            ciudad=hoja_activa.cell(row=fila,column=4).value if hoja_activa.cell(row=fila,column=4).value!=None else None
            lista_ciudades.append(ciudad)
            fila+=1
    
        lista_ciudades_mas_un_negocio=[]
        for ciudad in lista_ciudades:
            print(f"{ciudad}={lista_ciudades.count(ciudad)}")
        for ciudad in lista_ciudades:
            if lista_ciudades.count(ciudad)>1:
                if ciudad not in lista_ciudades_mas_un_negocio:
                    lista_ciudades_mas_un_negocio.append(ciudad)         
        return lista_ciudades_mas_un_negocio
    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")