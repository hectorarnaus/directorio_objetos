import re, html
from random import choice,randint
from ficheros_datos.keywords import *
from ficheros_datos.datos_población import *
from crea_elementos_web import *
from ficheros_datos.constantes_configuracion import *
from funciones_excel import ultima_fila_real

import openpyxl 


def spinner(s):
     
    while True:
        s, n = re.subn('{([^{}]*)}',
                    lambda m: choice(m.group(1).split("|")),
                    s)
        if n == 0: break
    return s.strip()

def obten_texto_H1(provincia):
    try:
        datos=openpyxl.load_workbook(excel_provincias)
        hoja_activa = datos.active
        fila=1
        while fila<ultima_fila_real(hoja_activa):
            if hoja_activa.cell(row=fila,column=1).value==provincia:
                return hoja_activa.cell(row=fila,column=6).value
            fila+=1

    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")
    return ""


def obten_texto_cuerpo_localidad(localidad):
    try:
        datos=openpyxl.load_workbook(excel_municipios)
        hoja_activa = datos.active
        fila=1
        while fila<ultima_fila_real(hoja_activa):
            if hoja_activa.cell(row=fila,column=1).value==localidad:
                return hoja_activa.cell(row=fila,column=5).value
            fila+=1

    except FileNotFoundError:
        print("Error: Archivo no encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")
    return ""
def maqueta_texto_cuerpo_localidad(texto):
    parrafos=texto.split("</p><p>")
    res=""
    for parrafo in parrafos:
        parrafo=parrafo.replace("<p>","")
        parrafo=parrafo.replace("</p>","")
        parrafo=parrafo.strip()
        res+=f'{crea_parrafo(parrafo)}'
    return res
def obten_texto_H1_old(provincia):
    with open('plantillas_textos/H1_provincia.txt', 'r') as file :
        texto_base = file.read()
        texto=spinner(texto_base)
        texto=texto.replace("*Provincia*",provincia)
        cantidad_keywords=texto.find("#keyword#")
        i=0
        keywords_usadas=[]
        while i<cantidad_keywords:
            nueva_keyword=randint(0,len(keywords_h1)-1)
            if nueva_keyword not in keywords_usadas:
                keywords_usadas.append(nueva_keyword)
                texto=texto.replace("#keyword#",keywords_h1[nueva_keyword],1)      
            i+=1
        return texto
    
def obten_texto_H2(fichero,lista_keywords,provincia):
    with open(f'plantillas_textos/{fichero}', 'r') as file :
        texto_base = file.read()
        texto=spinner(texto_base)
        texto=texto.replace("*Provincia*",provincia)
        cantidad_keywords=texto.find("#keyword#")
        i=0
        keywords_usadas=[]
        while i<cantidad_keywords:
            nueva_keyword=lista_keywords[randint(0,len(lista_keywords)-1)]
            if nueva_keyword not in keywords_usadas:
                keywords_usadas.append(nueva_keyword)
                texto=texto.replace("#keyword#",nueva_keyword,1)      
            i+=1
        return texto   


def obten_texto_provincia(provincia):
    with open('provincia.txt', 'r') as file :
        texto_base = file.read()
        texto=spinner(texto_base)
        texto=texto.replace("*Provincia*",provincia)
        return texto
    
def obten_texto_municipio(municipio):
    with open('municipio.txt', 'r') as file :
        texto_base = file.read()
        texto=spinner(texto_base)
        texto=texto.replace("*Municipio*",municipio)
        return texto
    
def crea_texto_ciudad(ciudad):
    res=""
    tipo_keyword=["transaccional","informacional","maquinaria","sector","urgencia"]
    if ciudad in prioridad_maxima:
        cantidad_keywords=5
    elif ciudad in prioridad_alta:
        cantidad_keywords=4
    elif ciudad in prioridad_media:
        cantidad_keywords=3
    else:
        cantidad_keywords=2
        
    keywords_usadas=[]
    i=0
    while i<cantidad_keywords:
        nueva_keyword=randint(0,len(tipo_keyword)-1)
        if nueva_keyword not in keywords_usadas:
            keywords_usadas.append(nueva_keyword)
            if nueva_keyword==0:
                res+=f'{crea_heading(f"Encuentra las mejores {tipo_negocio.lower()} en {ciudad}",2)}'
                res+=f'{crea_parrafo(obten_texto_H2("H2_transaccional.txt",keywords_transaccionales,ciudad))}'
            elif nueva_keyword==1:
                res+=f'{crea_heading(f"Guía para {tipo_negocio.lower()} en {ciudad}",2)}'
                res+=f'{crea_parrafo(obten_texto_H2("H2_informacional.txt",keywords_informacionales,ciudad))}'
            elif nueva_keyword==2:
                res+=f'{crea_heading(f"Tipos de maquinaria disponible en {ciudad}",2)}'
                res+=f'{crea_parrafo(obten_texto_H2("H2_tipo_maquinaria.txt",keywords_tipo_maquinaria,ciudad))}'
            elif nueva_keyword==3:
                res+=f'{crea_heading(f"Maquinaria especializada según tu tipo de proyecto en {ciudad}",2)}'
                res+=f'{crea_parrafo(obten_texto_H2("H2_proyecto_sector.txt",keywords_proyecto_sector,ciudad))}'
            elif nueva_keyword==4:
                res+=f'{crea_heading(f"Alquiler flexible de maquinaria y servicios urgentes en {ciudad}",2)}'
                res+=f'{crea_parrafo(obten_texto_H2('H2_urgencia_flexibilidad.txt',keywords_urgencia_flexibilidad,ciudad))}'
            i+=1
   
    return res
        
        


